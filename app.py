import streamlit as st
import pandas as pd
import re
from io import BytesIO
from tqdm import tqdm

st.set_page_config(layout="wide")

st.title("📊 BBH Tracker Automation Tool")

# ======================================
# FILE UPLOAD (UI)
# ======================================

tracker_file = st.file_uploader("Upload Tracker File", type=["xlsx"])
raw_file = st.file_uploader("Upload Raw BH File", type=["xlsx"])
day_file = st.file_uploader("Upload Day File", type=["xlsx"])
activation_file = st.file_uploader("Upload Activation File", type=["xlsx"])


# ======================================
# YOUR MAIN LOGIC (PUT YOUR CODE HERE)
# ======================================

def run_bbh(tracker, raw, day, activation):

    # 👉 MOVE YOUR FULL EXISTING LOGIC HERE
    # Example (keep yours, this is demo):


    # ======================================
    # FILE PATHS
    # ======================================



    # ======================================
    # LOAD FILES
    # ======================================


    # ======================================
    # DATE EXTRACTION
    # ======================================

    raw["Date"] = pd.to_datetime(raw["Period start time"]).dt.strftime("%d-%b-%y")
    day["Date"] = pd.to_datetime(day["Period start time"]).dt.strftime("%d-%b-%y")

    # ======================================
    # BAND / SECTOR EXTRACTION
    # ======================================

    def extract_band_sector(cell):

        if pd.isna(cell):
            return "Unknown","S0"

        cell = str(cell).upper()

        patterns = {
            r'4LSN(\d)$':"L2600-F2",
            r'4LN(\d)$':"L2600-F2",
            r'4LS(\d)$':"L2600-F1",
            r'4L(\d)$':"L2600-F1",
            r'4U(\d)$':"L2100",
            r'4G(\d)$':"L1800",
            r'4C(\d)$':"L800"
        }

        for p,b in patterns.items():

            m = re.search(p,cell)

            if m:
                return b,"S"+m.group(1)

        return "Other","S0"


    # ======================================
    # OVERRIDE
    # ======================================

    def override_band_sector(lncel,band,sec):

        corrections = {      
            "CHIBABVIRES4G4": ("L2100", "Sector-4"),"CHIBABVIRES4G5": ("L2100", "Sector-5"),"CHIBABVIRES4G6": ("L2100", "Sector-6"),
            "CHILAMBULAL4G4": ("L2100", "Sector-4"),"CHILAMBULAL4G5": ("L2100", "Sector-5"),"CHILAMBULAL4G6": ("L2100", "Sector-6"),
            "MCHINJIBOMA4G4": ("L2100", "Sector-4"),"MCHINJIBOMA4G5": ("L2100", "Sector-5"),"MCHINJIBOMA4G6": ("L2100", "Sector-6"),
            "CUNIMA4G4": ("L2100", "Sector-4"),"CUNIMA4G5": ("L2100", "Sector-5"),"CUNIMA4G6": ("L2100", "Sector-6"),
            "KALONGA4G4": ("L2100", "Sector-4"),"KALONGA4G5": ("L2100", "Sector-5"),"KALONGA4G6": ("L2100", "Sector-6"),
            "CHIMWETAPRIMARY4G4": ("L2100", "Sector-4"),"CHIMWETAPRIMARY4G5": ("L2100", "Sector-5"),"CHIMWETAPRIMARY4G6": ("L2100", "Sector-6"),
            "MPONELARES4G4": ("L2100", "Sector-4"),"MPONELARES4G5": ("L2100", "Sector-5"),"MPONELARES4G6": ("L2100", "Sector-6"),
            "LINDE4G4": ("L2100", "Sector-4"),"LINDE4G5": ("L2100", "Sector-5"),"LINDE4G6": ("L2100", "Sector-6"),
            "AREA49ZEBRA4G4": ("L2100", "Sector-4"),"AREA49ZEBRA4G5": ("L2100", "Sector-5"),"AREA49ZEBRA4G6": ("L2100", "Sector-6"),
            "NKHOTAKOTAHOS4G4": ("L2100", "Sector-4"),"NKHOTAKOTAHOS4G5": ("L2100", "Sector-5"),"NKHOTAKOTAHOS4G6": ("L2100", "Sector-6"),
            "AREA18FILLING4G4": ("L2100", "Sector-4"),"AREA18FILLING4G5": ("L2100", "Sector-5"),"AREA18FILLING4G6": ("L2100", "Sector-6"),
            "NEWGULLIVER4G4": ("L2100", "Sector-4"),"NEWGULLIVER4G5": ("L2100", "Sector-5"),"NEWGULLIVER4G6": ("L2100", "Sector-6"),
            "BAGHDAD24G4": ("L2100", "Sector-4"),"BAGHDAD24G5": ("L2100", "Sector-5"),"BAGHDAD24G6": ("L2100", "Sector-6"),
            "CHIMKUSA4G4": ("L2100", "Sector-4"),"CHIMKUSA4G5": ("L2100", "Sector-5"),"CHIMKUSA4G6": ("L2100", "Sector-6"),
            "GULLIVERPUBS4G4": ("L2100", "Sector-4"),"GULLIVERPUBS4G5": ("L2100", "Sector-5"),"GULLIVERPUBS4G6": ("L2100", "Sector-6"),
            "MCHENGAUTUBA4G4": ("L2100", "Sector-4"),"MCHENGAUTUBA4G5": ("L2100", "Sector-5"),"MCHENGAUTUBA4G6": ("L2100", "Sector-6"),
        }


        lncel = str(lncel).upper()

        if lncel in corrections:
            return corrections[lncel]

        return band,sec


    # ======================================
    # APPLY BAND / SEC
    # ======================================

    raw[["Band","Sec"]] = raw["LNCEL name"].apply(
        lambda x: pd.Series(extract_band_sector(x))
    )

    raw[["Band","Sec"]] = raw.apply(
        lambda r: override_band_sector(
            r["LNCEL name"],
            r["Band"],
            r["Sec"]
        ),
        axis=1,
        result_type="expand"
    )

    raw["SecName"] = raw["LNBTS name"]+"_"+raw["Sec"]

    # ======================================
    # SAME FOR DAY FILE
    # ======================================

    day[["Band","Sec"]] = day["LNCEL name"].apply(
        lambda x: pd.Series(extract_band_sector(x))
    )

    day[["Band","Sec"]] = day.apply(
        lambda r: override_band_sector(
            r["LNCEL name"],
            r["Band"],
            r["Sec"]
        ),
        axis=1,
        result_type="expand"
    )

    day["SecName"] = day["LNBTS name"]+"_"+day["Sec"]

    # ======================================
    # RENAME DAY KPIs
    # ======================================

    day = day.rename(columns={
    "Total LTE data volume, DL + UL":"Total LTE data volume, DL + UL(Daily)",
    "Cell Avail excl BLU":"Cell Avail excl BLU(Daily)"
    })

    # ======================================
    # KPI LIST
    # ======================================

    raw_kpis = [
        "Total E-UTRAN RRC conn stp SR2","Total E-UTRAN RRC conn stp SR",
        "Intra eNB HO SR total","E-UTRAN E-RAB stp SR","E-RAB DR RAN",
        "Intra eNB HO SR","E-UTRAN Intra-Freq HO SR",
        "inter eNB E-UTRAN HO SR X2","Avg RRC conn UE","Average CQI",
        "Data RB stp SR","Avg UE distance",
        "UserDownlinkAverageThroughput",
        "RRC_CONN_UE_MAX (M8001C200)",
        "User Uplink Avg Throughput152","RSSI_PUCCH_AVG (M8005C2)",
        "Avg RSSI for PUSCH","SINR_PUCCH_AVG (M8005C92)",
        "SINR_PUSCH_AVG (M8005C95)","RACH Stp Completion SR",
        "Perc DL PRB Util","Init Contx stp SR for CSFB",
        "% MIMO RI 2","% MIMO RI 1","% MIMO RI 4",
        "PDCP SDU Volume, DL","PDCP SDU Volume, UL","Total LTE data volume, DL + UL","Cell Avail excl BLU"
    ]

    day_kpis = [
    "Total LTE data volume, DL + UL(Daily)",
    "Cell Avail excl BLU(Daily)"
    ]

    # ======================================
    # FILTER ACTIVATED SITES
    # ======================================

    sites = activation[
    activation["Activation Date"].notna()
    ]["LNBTS name"].unique()

    #raw = raw[raw["LNBTS name"].isin(sites)]
    #day = day[day["LNBTS name"].isin(sites)]

    # ======================================
    # RAW LONG FORMAT
    # ======================================

    records = []

    for kpi in tqdm(raw_kpis, desc="Processing BBH KPIs"):

        if kpi in raw.columns:

            temp = raw[
            ["LNBTS name","LNCEL name","Band","Sec","SecName","Date",kpi]
            ].copy()

            temp["KPI Name"] = kpi
            temp = temp.rename(columns={kpi:"Value"})

            records.append(temp)

    raw_long = pd.concat(records)
    raw_long = raw_long.drop_duplicates(subset=["LNBTS name","LNCEL name","KPI Name","Date"])

    # ======================================
    # DAY LONG FORMAT
    # ======================================

    records = []

    for kpi in tqdm(day_kpis, desc="Processing DAY KPIs"):

        if kpi in day.columns:

            temp = day[
            ["LNBTS name","LNCEL name","Band","Sec","SecName","Date",kpi]
            ].copy()

            temp["KPI Name"] = kpi
            temp = temp.rename(columns={kpi:"Value"})

            records.append(temp)

    day_long = pd.concat(records)
    day_long = day_long.drop_duplicates(subset=["LNBTS name","LNCEL name","KPI Name","Date"])

    # ======================================
    # COMBINE BOTH
    # ======================================

    combined = pd.concat([raw_long,day_long])

    # ======================================
    # CREATE PIVOT
    # ======================================

    pivot = combined.pivot_table(

    index=[
    "LNBTS name",
    "LNCEL name",
    "Band",
    "Sec",
    "SecName",
    "KPI Name"
    ],

    columns="Date",

    values="Value",

    aggfunc="mean"

    ).reset_index()

    # ======================================
    # ADD ACTIVATION DATA
    # ======================================

    pivot = pivot.merge(

    activation[["LNBTS name","Activation Date","Remark"]],

    on="LNBTS name",

    how="left"

    )

    pivot = pivot.drop(columns=[
        "Activation Date",
        "Remark"
    ], errors="ignore")

    # ======================================
    # CHECK MISSING CELLS
    # ======================================

    tracker_keys=set(

    zip(

    tracker["LNBTS name"],

    tracker["LNCEL name"],

    tracker["KPI Name"]

    )

    )

    new_rows=[]

    for _,row in tqdm(pivot.iterrows(), total=len(pivot), desc="Checking new rows"):

        key=(row["LNBTS name"],row["LNCEL name"],row["KPI Name"])

        if key not in tracker_keys:

            new_rows.append(row)

    if new_rows:

        new_rows_df = pd.DataFrame(new_rows)

        # ==================================
        # ADD ACTIVATION DATE + REMARK HERE
        # ==================================

        new_rows_df = pd.DataFrame(new_rows)
        
        # ensure Band/Sec/SecName go to tracker columns
        new_rows_df["Band"] = new_rows_df["Band"]
        new_rows_df["Sec"] = new_rows_df["Sec"]
        new_rows_df["SecName"] = new_rows_df["SecName"]
        
        activation_lookup = activation.drop_duplicates("LNBTS name").set_index("LNBTS name")["Activation Date"]
        
        new_rows_df["Activation Date"] = new_rows_df["LNBTS name"].map(activation_lookup)
        
        new_rows_df["Remark"] = new_rows_df["Activation Date"].apply(
            lambda x: "Consider" if pd.notna(x) else ""
        )
        
        tracker = pd.concat([tracker,new_rows_df],ignore_index=True)
        
        activation_lookup = activation.drop_duplicates("LNBTS name").set_index("LNBTS name")["Activation Date"]

        tracker["Activation Date"] = tracker["LNBTS name"].map(activation_lookup).fillna(tracker["Activation Date"])
        
        tracker["Remark"] = tracker.apply(
            lambda x: "Consider" if pd.notna(x["Activation Date"]) else x["Remark"],
            axis=1
        )
        date_cols = tracker.columns[tracker.columns.get_loc("Activation Date")+1:]
        tracker[date_cols] = tracker[date_cols].round(2)


    # ======================================
    # MERGE VALUES
    # ======================================

    pivot_merge = pivot.drop(
        columns=["Band","Sec","SecName","Activation Date","Remark"],
        errors="ignore"
    )

    tracker = tracker.merge(

        pivot_merge,

        on=["LNBTS name","LNCEL name","KPI Name"],

        how="left",

        suffixes=("","_new")

    )
    # ======================================
    # UPDATE DATE COLUMNS
    # ======================================

    date_cols = tracker.columns[tracker.columns.get_loc("Activation Date")+1:]

    sum_kpis = [
        "Total LTE data volume, DL + UL(Daily)",
        "Total LTE data volume, DL + UL",
        "PDCP SDU Volume, DL",
        "PDCP SDU Volume, UL"
    ]

    avg_kpis = [
        "UserDownlinkAverageThroughput",
        "Perc DL PRB Util",
        "Avg RRC conn UE",
        "Average CQI",
        "Avg UE distance",
        "Total E-UTRAN RRC conn stp SR",
        "E-RAB DR RAN",
        "E-UTRAN E-RAB stp SR",
        "RACH Stp Completion SR",
        "Cell Avail excl BLU(Daily)",
        "Init Contx stp SR for CSFB",
        "E-UTRAN Intra-Freq HO SR",
        "inter eNB E-UTRAN HO SR X2",
        "Intra eNB HO SR total"
    ]

    for col in date_cols:

        new=str(col)+"_new"

        if new in tracker.columns:

            tracker[col]=tracker[new].combine_first(tracker.get(col))

            tracker.drop(columns=[new],inplace=True)
            
            
    # ======================================
    # REMOVE DUPLICATE CELL-KPI ROWS
    # ======================================

    print("Checking duplicates in tracker...")

    tracker["CELL_KPI_KEY"] = tracker["LNCEL name"] + "_" + tracker["KPI Name"]

    before_rows = len(tracker)

    tracker = tracker.drop_duplicates(subset=["CELL_KPI_KEY"], keep="first")

    after_rows = len(tracker)

    print("Duplicate rows removed:", before_rows - after_rows)

    tracker.drop(columns=["CELL_KPI_KEY"], inplace=True)
            
    # =================================================
    # SHEET 2 : GLOBAL KPI SUMMARY (Simple)
    # =================================================

    # Copy tracker
    summary_df = tracker.copy()

    # Filter 'Consider' rows
    summary_df = tracker[tracker["Remark"]=="Consider"].copy()

    remove_kpis = [
        "Avg RSSI for PUSCH",
        "Cell Avail excl BLU",
        "Data RB stp SR",
        "E-UTRAN Intra-Freq HO SR",
        "Intra eNB HO SR",
        "RSSI_PUCCH_AVG (M8005C2)",
        "SINR_PUCCH_AVG (M8005C92)",
        "SINR_PUSCH_AVG (M8005C95)"
    ]

    summary_df = summary_df[~summary_df["KPI Name"].isin(remove_kpis)]

    # Dynamically get all date columns
    exclude_cols = ["LNBTS name","LNCEL name","Band","Sec","SecName","KPI Name","Activation Date","Remark"]
    date_cols = [c for c in tracker.columns if c not in exclude_cols]

    payload_kpis = ["Total LTE data volume, DL + UL","Total LTE data volume, DL + UL(Daily)","Avg RRC conn UE","PDCP SDU Volume, DL","PDCP SDU Volume, UL"]

    summary_rows = []

    for kpi, grp in summary_df.groupby("KPI Name"):
        row = {"KPI Name": kpi}
        for d in date_cols:
            if kpi in payload_kpis:
                # SUM over all 'Consider' rows, ignoring NaN
                row[d] = grp[d].sum(skipna=True)
            else:
                # MEAN over all 'Consider' rows, ignoring NaN
                row[d] = grp[d].mean(skipna=True)
        summary_rows.append(row)

    sheet2 = pd.DataFrame(summary_rows)

    # ===============================
    # ADD CUSTOM KPIs TO SUMMARY
    # ===============================

    thr_df = summary_df[summary_df["KPI Name"]=="UserDownlinkAverageThroughput"]
    prb_df = summary_df[summary_df["KPI Name"]=="Perc DL PRB Util"]

    total_cells = summary_df["LNCEL name"].nunique()

    # -----------------------------
    # TOTAL CELLS
    # -----------------------------
    row_total = {"KPI Name":"Total number of cells"}
    for d in date_cols:
        row_total[d] = total_cells

    # -----------------------------
    # CELLS <3 Mbps
    # -----------------------------
    row_3 = {"KPI Name":"LTE_DL-user_throughput_Mbps(<3Mbps)#cells Count"}

    for d in date_cols:
        temp = thr_df[thr_df[d] < 3000]
        row_3[d] = temp["LNCEL name"].nunique()

    # -----------------------------
    # CELLS <5 Mbps
    # -----------------------------
    row_5 = {"KPI Name":"LTE_DL-user_throughput_Mbps(<5Mbps)#cells Count"}

    for d in date_cols:
        temp = thr_df[thr_df[d] < 5000]
        row_5[d] = temp["LNCEL name"].nunique()

    # -----------------------------
    # CELLS >70% PRB
    # -----------------------------
    row_prb = {"KPI Name":"Count of cells >70% PRB"}

    for d in date_cols:
        row_prb[d] = (prb_df[d] > 70).sum()

    # -----------------------------
    # % >70% PRB
    # -----------------------------
    row_prb_pct = {"KPI Name":"% Of cells  >70% PRB"}

    for d in date_cols:
        row_prb_pct[d] = (row_prb[d] / total_cells) * 100

    # -----------------------------
    # % <3 Mbps
    # -----------------------------
    row_3_pct = {"KPI Name":"% Of cells , LTE_DL-user_throughput_Mbps(<3Mbps)"}

    for d in date_cols:
        row_3_pct[d] = (row_3[d] / total_cells) * 100

    # -----------------------------
    # % <5 Mbps
    # -----------------------------
    row_5_pct = {"KPI Name":"% Of cells , LTE_DL-user_throughput_Mbps(<5Mbps)"}

    for d in date_cols:
        row_5_pct[d] = (row_5[d] / total_cells) * 100

    # -----------------------------
    # APPEND ALL
    # -----------------------------
    sheet2 = pd.concat([
        sheet2,
        pd.DataFrame([
            row_total,
            row_3,
            row_5,
            row_prb_pct,
            row_3_pct,
            row_5_pct
        ])
    ], ignore_index=True)

    # Round numeric values
    sheet2[date_cols] = sheet2[date_cols].round(2)

    # Format date columns safely (only actual dates)
    formatted_cols = {}
    for c in date_cols:
        try:
            dt = pd.to_datetime(c)
            formatted_cols[c] = dt.strftime("%d-%b-%y")
        except:
            formatted_cols[c] = c

    sheet2.rename(columns=formatted_cols, inplace=True)

    # ======================================
    # SAVE
    # ======================================

    # ======================================
    # FUNCTION: PRE / POST / DELTA / %
    # ======================================
    def add_pre_post_delta(df):

        import pandas as pd

        exclude_cols = ["LNBTS name","LNCEL name","Band","Sec","SecName",
                        "KPI Name","Activation Date","Remark","KPI"]

        date_cols = [c for c in df.columns if c not in exclude_cols]

        # Map columns to datetime
        date_map = {}
        for c in date_cols:
            try:
                date_map[c] = pd.to_datetime(c, format="%d-%b-%y")
            except:
                continue

        # Sort dates
        sorted_dates = sorted(date_map.items(), key=lambda x: x[1])

        if len(sorted_dates) < 3:
            return df

        # -------------------------
        # POST → Latest 3 days
        # -------------------------
        post_cols = [c for c,_ in sorted_dates[-3:]]
        post_days = [d.day_name() for _,d in sorted_dates[-3:]]

        # -------------------------
        # PRE → 19–25 Jan window
        # -------------------------
        pre_start = pd.to_datetime("2026-01-19")
        pre_end = pd.to_datetime("2026-01-25")

        pre_cols = []
        for day in post_days:
            for c,d in sorted_dates:
                if pre_start <= d <= pre_end and d.day_name() == day:
                    pre_cols.append(c)
                    break

        if len(pre_cols) == 0:
            return df

        # -------------------------
        # CALCULATIONS
        # -------------------------
        df["Pre (19–25 Jan)"] = df[pre_cols].mean(axis=1)
        df["Post (Last 3 Days)"] = df[post_cols].mean(axis=1)

        df["Delta"] = df["Post (Last 3 Days)"] - df["Pre (19–25 Jan)"]

        df["% Increase"] = (df["Delta"] / df["Pre (19–25 Jan)"]) * 100

        # Handle divide by zero
        df["% Increase"] = df["% Increase"].replace([float('inf'), -float('inf')], 0)

        return df



    # ===============================
    # SHEET 3 : BAND SUMMARY
    # ===============================

    date_cols = tracker.columns[tracker.columns.get_loc("Activation Date")+1:]

    bands = ["L800","L1800","L2100","L26C1","L26C2"]

    sheet3_rows = []

    for band in tqdm(bands, desc="Processing Band Summary"):

        band_df = tracker[
            (tracker["Band"] == band) &
            (tracker["Remark"] == "Consider")
        ]

        if band_df.empty:
            continue

        # -----------------------------
        # TOTAL CELLS
        # -----------------------------
        total_cells = band_df["LNCEL name"].nunique()

        # -----------------------------
        # SUM KPIs
        # -----------------------------
        for kpi in sum_kpis:

            kpi_df = band_df[band_df["KPI Name"]==kpi]

            if kpi_df.empty:
                continue

            valid_dates = [c for c in date_cols if c in kpi_df.columns]

            values = kpi_df[valid_dates].sum()

            row = {"KPI":kpi,"Band":band}
            row.update(values)

            sheet3_rows.append(row)

        # -----------------------------
        # AVERAGE KPIs
        # -----------------------------
        for kpi in avg_kpis:
        
            kpi_df = band_df[band_df["KPI Name"]==kpi]
        
            if kpi_df.empty:
                continue
        
            valid_dates = [c for c in date_cols if c in kpi_df.columns]
        
            values = kpi_df[valid_dates].mean()
        
            row = {"KPI":kpi,"Band":band}
            row.update(values)
        
            sheet3_rows.append(row)

        # -----------------------------
        # COUNT CELLS >70% PRB
        # -----------------------------
        prb_df = band_df[band_df["KPI Name"]=="Perc DL PRB Util"]

        valid_dates = [c for c in date_cols if c in prb_df.columns]
        prb_count = (prb_df[valid_dates] > 70).sum()

        row = {"KPI":"Count of cells >70% PRB","Band":band}
        row.update(prb_count)

        sheet3_rows.append(row)

        # -----------------------------
        # THROUGHPUT <3 Mbps
        # -----------------------------
        thr_df = band_df[band_df["KPI Name"]=="UserDownlinkAverageThroughput"]
        
        valid_dates = [c for c in date_cols if c in thr_df.columns]
        
        row = {"KPI":"LTE_DL-user_throughput_Mbps(<3Mbps)#cells Count","Band":band}
        
        for d in valid_dates:
            
            temp = thr_df[thr_df[d] < 3000]
            
            # count unique cells
            row[d] = temp["LNCEL name"].nunique()
        
        sheet3_rows.append(row)

        # -----------------------------
        # THROUGHPUT <5 Mbps
        # -----------------------------
        valid_dates = [c for c in date_cols if c in thr_df.columns]
        thr5 = (thr_df[valid_dates] < 5).sum()

        row = {"KPI":"LTE_DL-user_throughput_Mbps(<5Mbps)#cells Count","Band":band}
        
        for d in valid_dates:
            
            temp = thr_df[thr_df[d] < 5000]
            
            row[d] = temp["LNCEL name"].nunique()
        
        sheet3_rows.append(row)

        # -----------------------------
        # TOTAL CELLS
        # -----------------------------
        row = {"KPI":"Total number of cells","Band":band}

        for d in date_cols:
            row[d] = total_cells

        sheet3_rows.append(row)

    sheet3 = pd.DataFrame(sheet3_rows)

    # ===============================
    # % Traffic Distribution
    # ===============================

    traffic = sheet3[sheet3["KPI"]=="Total LTE data volume, DL + UL(Daily)"]

    traffic_pct = traffic.copy()
    traffic_pct["KPI"] = "% Traffic Distribution"

    for d in date_cols:
        total = traffic[d].sum()
        traffic_pct[d] = (traffic[d] / total) * 100

    sheet3 = pd.concat([sheet3,traffic_pct],ignore_index=True)

    # ===============================
    # % User Distribution
    # ===============================
    users = sheet3[sheet3["KPI"]=="Avg RRC conn UE"]

    users_pct = users.copy()
    users_pct["KPI"] = "% User Distribution"

    for d in date_cols:
        total = users[d].sum()
        users_pct[d] = (users[d] / total) * 100

    sheet3 = pd.concat([sheet3,users_pct],ignore_index=True)

    # ===============================
    #% Cells >70% PRB
    # ===============================

    cells70 = sheet3[sheet3["KPI"]=="Count of cells >70% PRB"]
    cells_total = sheet3[sheet3["KPI"]=="Total number of cells"]

    pct70 = cells70.copy()
    pct70["KPI"] = "% Of cells  >70% PRB"

    for d in date_cols:
        pct70[d] = (cells70[d].values / cells_total[d].values) * 100

    sheet3 = pd.concat([sheet3,pct70],ignore_index=True)

    # ===============================
    #%  Cells Throughput <3 Mbps & <5
    # ===============================

    cells3 = sheet3[sheet3["KPI"]=="LTE_DL-user_throughput_Mbps(<3Mbps)#cells Count"]
    cells_total = sheet3[sheet3["KPI"]=="Total number of cells"]

    pct3 = cells3.copy()
    pct3["KPI"] = "% Of cells , LTE_DL-user_throughput_Mbps(<3Mbps)"

    for d in date_cols:
        
        pct3[d] = (
            cells3[d].astype(float).values /
            cells_total[d].astype(float).values
        ) * 100

    sheet3 = pd.concat([sheet3,pct3],ignore_index=True)

    cells5 = sheet3[sheet3["KPI"]=="LTE_DL-user_throughput_Mbps(<5Mbps)#cells Count"]

    pct5 = cells5.copy()
    pct5["KPI"] = "% Of cells , LTE_DL-user_throughput_Mbps(<5Mbps)"

    for d in date_cols:
        pct5[d] = (cells5[d].values / cells_total[d].values) * 100

    sheet3 = pd.concat([sheet3,pct5],ignore_index=True)




    #sheet3 = pd.DataFrame(sheet3_rows)
    sheet3[date_cols] = sheet3[date_cols].round(2)

    # ======================================
    # FORMAT DATE COLUMNS (ALL SHEETS)
    # ======================================

    exclude_cols = ["LNBTS name","LNCEL name","Band","Sec","SecName","KPI Name","Activation Date","Remark"]

    date_cols = [c for c in tracker.columns if c not in exclude_cols]

    formatted_cols = {}

    for c in date_cols:
        try:
            dt = pd.to_datetime(c)
            formatted_cols[c] = dt.strftime("%d-%b-%y")
        except:
            formatted_cols[c] = c

    # Apply to all sheets
    tracker.rename(columns=formatted_cols, inplace=True)
    sheet3.rename(columns=formatted_cols, inplace=True)

    # ======================================
    # APPLY PRE / POST / DELTA / %
    # ======================================

    tracker = add_pre_post_delta(tracker)
    sheet2 = add_pre_post_delta(sheet2)
    sheet3 = add_pre_post_delta(sheet3)

    # Round values
    for df in [tracker, sheet2, sheet3]:
        for col in ["Pre (19–25 Jan)", "Post (Last 3 Days)", "Delta", "% Increase"]:
            if col in df.columns:
                df[col] = df[col].round(2)





    from openpyxl.styles import PatternFill


    output = BytesIO()
    with pd.ExcelWriter(output,engine="openpyxl") as writer:

        tracker.to_excel(writer,sheet_name="BBH Tracker",index=False)
        sheet2.to_excel(writer,sheet_name="Summary",index=False)
        sheet3.to_excel(writer, sheet_name="Band Summary", index=False)


        # ======================================
        # APPLY COLOR FORMATTING (BBH Tracker)
        # ======================================

        ws = writer.sheets["BBH Tracker"]

        # Colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # KPI Rules
        kpi_rules = {
            "UserDownlinkAverageThroughput": lambda x: x > 5000,
            "Total E-UTRAN RRC conn stp SR": lambda x: x > 99.5,
            "E-RAB DR RAN": lambda x: x < 0.3,
            "E-UTRAN E-RAB stp SR": lambda x: x > 99.5,
            "RACH Stp Completion SR": lambda x: x > 99.5,
            "Cell Avail excl BLU(Daily)": lambda x: x > 99,
            "Init Contx stp SR for CSFB": lambda x: x > 99.5,
            "inter eNB E-UTRAN HO SR X2": lambda x: x > 95,
            "Intra eNB HO SR total": lambda x: x > 97
        }

        # Get headers
        headers = [cell.value for cell in ws[1]]

        kpi_col = headers.index("KPI Name")
        date_start_col = headers.index("Activation Date") + 1

        # Loop rows
        for row in ws.iter_rows(min_row=2):

            kpi = row[kpi_col].value

            for cell in row[date_start_col:]:

                try:
                    val = float(cell.value)
                except:
                    continue

                # PRB special logic
                if kpi == "Perc DL PRB Util":
                    if val > 90:
                        cell.fill = red_fill
                    elif val > 70:
                        cell.fill = green_fill

                # Other KPIs
                elif kpi in kpi_rules:
                    if kpi_rules[kpi](val):
                        cell.fill = green_fill
            
                cell.number_format = '0.00'
    return output

    print("BBH Tracker Updated Successfully")


# ======================================
# RUN BUTTON
# ======================================

if st.button("🚀 Run BBH Process"):

    if tracker_file and raw_file and day_file and activation_file:

        # READ FILES
        tracker = pd.read_excel(tracker_file)
        raw = pd.read_excel(raw_file)
        day = pd.read_excel(day_file)
        activation = pd.read_excel(activation_file)

        st.info("Processing...")

        # RUN YOUR FUNCTION
        result = run_bbh(tracker, raw, day, activation)

        st.success("✅ Done!")

        # DOWNLOAD
        output = BytesIO()
        result = run_bbh(tracker, raw, day, activation)

        st.download_button(
            label="📥 Download Output",
            data=result.getvalue(),
            file_name="BBH_Output.xlsx"
        )

    else:
        st.warning("⚠️ Upload all files first")
